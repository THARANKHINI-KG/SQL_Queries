from openpyxl import load_workbook, Workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet.append(["thara", 21])
sheet.append(["alice", 20])
wb.save(r"C:\Users\thara\OneDrive\Desktop\msg.xlsx")

wb = load_workbook(r"C:\Users\thara\OneDrive\Desktop\msg.xlsx")
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)