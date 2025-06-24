# file handling

# excel
# creating the excel, writing data
from openpyxl import Workbook
import os
wb1 = Workbook()
ws1 = wb1.active
for i in range(1,21):
    ws1.append([f"Row {i}", "*"*i])
folder = r"G:\THARA\hexaware\file handling"
file = "book.xlsx"
full_path = os.path.join(folder,file)
wb1.save(full_path)
print(f"File saved in {full_path}")

# reading data
from openpyxl import load_workbook
wb2 = load_workbook(r"G:\THARA\hexaware\file handling\book.xlsx", read_only=True)
ws2 = wb2.active
for row in ws2.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)

# writing data
from openpyxl import Workbook
wb3 = Workbook(write_only=True)
ws3 = wb3.create_sheet()
for i in range(1, 21):
    ws3.append([f"Row {i}", "*"*i])
folder = r"G:\THARA\hexaware\file handling"
file = "book1.xlsx"
full_path = os.path.join(folder,file)
wb3.save(full_path)
print(f"file saved in {full_path}")